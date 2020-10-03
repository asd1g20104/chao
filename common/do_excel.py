# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from chao.work.project.common import project_path
from chao.work.project.common.peizhiwenjian import ReadConfig




file_name = '../test_cases/test_api.xlsx'



class DoExcel:
    '''完成测试数据的读取'''


    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name



    def read_data(self,peizhi):
        '''从Excel读取数据'''

        case_id=ReadConfig(project_path.conf_path).get_data(peizhi,'case_id')
        wb=load_workbook(self.file_name)

        sheet=wb[self.sheet_name]


        test_data=[]
        for i in range(2,sheet.max_row+1):
            tel = self.get_tel()
            row_data={}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            if sheet.cell(i,6).value.find('tel')!=-1:#找tel找不到就返回-1
                row_data['Params'] = sheet.cell(i, 6).value.replace('tel',str(tel))#替换值,replace是字符串之间的替换。需要转为字符串，字符串 find tel replace tel表单里面的电话号码
                self.update_tel(int(tel)+1)
            else:
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['Sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            test_data.append(row_data)


        wb.save(self.file_name)
        wb.close()
        final_data=[]#空列表,存储最终的数据
        if case_id=='all':
             final_data=test_data
        else:
            for i in case_id:#遍历case_id里面的值
               final_data.append(test_data[i-1])
        return final_data

    def get_tel(self):
        '''获取Excel里面的tel值'''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        wb.close()
        return sheet.cell(1, 2).value  # 返回电话号码的值

    def update_tel(self,new_tel):
        '''写回电话号码'''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        sheet.cell(1,2,new_tel)
        wb.save(self.file_name)
        wb.close()


    @staticmethod
    def write_back(file_name,sheet_name,row,col,value):
        wb=load_workbook(file_name)

        # sheet=wb.get_sheet_by_name(sheet_name)
        sheet=wb[sheet_name]
        sheet.cell(row,col).value=value
        wb.save(file_name)
        wb.close




if __name__ == '__main__':
    sheet_name='register'
    test_data=DoExcel(project_path.case_path,sheet_name).read_data('test_recharge')
    print(test_data)

